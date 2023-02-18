from .conf import *

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Union


def find_column(worksheet: Worksheet, query_string="Area") -> Union[int, str]:
    """find column by title"""
    header = worksheet[1]
    guess = query_string.lower()

    for cell in header:
        if guess in cell.value.lower():  # TODO: add string similarity check
            print(
                f"""query_string {cell.value} found at column {cell.col_idx}
                    - with letter {cell.column_letter}
                """
            )
            return cell


def get_rows_data(worksheet):
    data = []

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):  # skip header
        data.append([cell for cell in row])
    return data


# function to get all data from a column
def get_column_data(worksheet, column_idx=None):
    if column_idx is None:
        column_idx = find_column(worksheet, query_string="area").col_idx

    data = []

    # alternatively, you can use list(ws.columns)[2] to get the third column
    for col in worksheet.iter_cols(
        min_row=2,
        min_col=column_idx,
        max_col=column_idx,
        values_only=True,
    ):
        for cell in col:
            data.append(cell)
    return data


def insert_area_columns(worksheet, idx=None, columns_number=2):
    """
    if we try to insert in coloumn 4 and it's already there, it will shift the rest of the columns to the right
    so we need to insert area_en, area_ar after the col with title (area)
    """
    if idx is None:
        idx = find_column(worksheet).col_idx

    next_col = idx + 1

    for _ in range(columns_number):
        worksheet.insert_cols(idx=next_col)
        print(f"inserted column at {next_col}")


def set_header_data(worksheet, next_col=None):
    if next_col is None:
        next_col = find_column(worksheet).col_idx + 1

    worksheet.cell(row=1, column=next_col).value = area_header[0]
    worksheet.cell(row=1, column=next_col + 1).value = area_header[1]
    print(
        find_column(worksheet, query_string="area_en"),
        find_column(worksheet, query_string="area_ar"),
    )


def set_rows_data(worksheet, next_col):
    # TODO: get input from user to continue or not if not make them inter the data and store it in a file or use cachelib
    for i, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        row[next_col - 1].value = area_data[i - 2][0]

        row[next_col].value = area_data[i - 2][1]


def save_to_xlsx_file(workbook: Workbook, file_name="output.xlsx"):
    user_filename = input(
        f"""Saving to {file_name}?
        Press enter to continue... or type a new filename: """
    )

    workbook.save(user_filename or file_name)
